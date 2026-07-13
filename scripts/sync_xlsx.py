#!/usr/bin/env python3
"""Sync watchlist.json from 汇总持仓.xlsx — handles 7/13 东方AI 赴回"""
import openpyxl, json

XLSX_PATH = 'C:/Users/LEGION/Downloads/汇总持仓.xlsx'
WL_PATH = 'C:/Users/LEGION/Nutstore/1/daily-report/config/watchlist.json'

wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
ws = wb['持仓数据']

headers = [cell.value for cell in ws[1]]
rows = []
for row in ws.iter_rows(min_row=2, max_row=ws.max_row-1, values_only=True):
    rows.append(dict(zip(headers, row)))

# Find 7/13 赴回 transaction
ws_tr = wb['交易记录']
tr_headers = [cell.value for cell in ws_tr[1]]
sell_info = None
for row in ws_tr.iter_rows(min_row=2, values_only=True):
    d = dict(zip(tr_headers, row))
    if d.get('成交日期') == '2026-07-13' and d.get('代码') == '017811' and d.get('交易类别') == '赎回':
        sell_info = d
        break

sell_shares = sell_info['成交数量']  # 368.6
sell_price = sell_info['成交价格']   # 4.0357
sell_proceeds = sell_info['发生金额'] # 1480.12
sell_fee = sell_info['费用']          # 7.4378
net_proceeds = sell_proceeds - sell_fee

# Read existing watchlist
with open(WL_PATH, 'r', encoding='utf-8') as f:
    wl = json.load(f)

xlsx_by_code = {r['代码']: r for r in rows}
holdings_map = {h['code']: h for h in wl['holdings']}

# === Step 1: Update all funds from xlsx (pre-赴回 state for 东方AI) ===
for code, xd in xlsx_by_code.items():
    if code in holdings_map:
        h = holdings_map[code]
        h['market_value'] = xd['持有金额']
        h['hold_return_amount'] = xd['持有盈亏']
        h['hold_return_pct'] = xd['持有盈亏率']
        wp = xd['仓位占比']
        h['weight_pct'] = round(wp * 100, 2) if wp < 1 else wp
        h['shares'] = xd['持有数量']
        h['cost_basis'] = xd['单位成本']
        h['breakeven_pct'] = xd.get('回本涨幅') or 0
        h['xlsx_days_held'] = xd['持仓天数']
        h['return_1m'] = xd.get('近1月涨幅') or 0
        h['return_3m'] = xd.get('近3月涨幅') or 0
        h['return_6m'] = xd.get('近6月涨幅') or 0
        h['return_1y'] = xd.get('近1年涨幅') or 0
        h['xlsx_synced'] = '2026-07-13'
        h['xlsx_sector'] = xd['关联板块']

# === Step 2: Apply 7/13 赴回 to 东方AI ===
ai = holdings_map['017811']
pre_sell_shares = ai['shares']  # 737.2 (from xlsx, = 1105.8 - 1/3 already sold on 7/10)
remaining_shares = pre_sell_shares - sell_shares  # 737.2 - 368.6 = 368.6
ratio = remaining_shares / pre_sell_shares

ai['shares'] = remaining_shares
ai['market_value'] = round(ai['market_value'] * ratio, 2)
ai['hold_return_amount'] = round(ai['hold_return_amount'] * ratio, 2)
ai['hold_return_pct'] = ai['hold_return_pct']  # per-share metric unchanged
ai['rebalance_role'] = '核心锁本对象(7/10卖1/3+7/13赎回368.6份,剩368.6份≈3成仓位)'

# Special: 东方阿尔法
holdings_map['024424']['rebalance_role'] = '核心锁本对象(第一大仓位)'

# Handle missing funds (煤炭+消费)
missing_codes = set(holdings_map.keys()) - set(xlsx_by_code.keys())
for code in missing_codes:
    h = holdings_map[code]
    h['status'] = 'cleared'
    h['market_value'] = 0

# === Step 3: Calculate totals ===
active_funds = [h for h in wl['holdings'] if h['status'] == 'active']
total_holdings_mv = sum(h['market_value'] for h in active_funds)

old_cash = wl['meta']['cash']
new_cash = round(old_cash + net_proceeds, 2)
new_total = round(total_holdings_mv + new_cash, 2)

# Recalculate weights
for h in active_funds:
    h['weight_pct'] = round(h['market_value'] / new_total * 100, 2)

# === Step 4: Update meta ===
total_hold_pnl = sum(h['hold_return_amount'] for h in active_funds)
hold_cost = total_holdings_mv - total_hold_pnl
hold_pnl_pct = round(total_hold_pnl / hold_cost * 100, 4) if hold_cost else 0

wl['meta']['account_total_asset'] = new_total
wl['meta']['account_daily_return'] = 0
wl['meta']['account_daily_return_pct'] = 0
wl['meta']['account_hold_return'] = round(total_hold_pnl, 2)
wl['meta']['account_hold_return_pct'] = hold_pnl_pct
wl['meta']['account_cumulative_return'] = round(total_hold_pnl, 2)
wl['meta']['account_cumulative_return_pct'] = hold_pnl_pct
wl['meta']['cash'] = new_cash
wl['meta']['cash_pct'] = round(new_cash / new_total * 100, 2)
wl['meta']['realized_profit_today'] = round(net_proceeds, 2)
wl['meta']['snapshot_date'] = '2026-07-13T10:47'
wl['meta']['snapshot_source'] = 'xlsx同步(汇总持仓.xlsx) + 7/13赎回东方AI368.6份'
wl['meta']['updated_at'] = '2026-07-13T10:47:00+08:00'
wl['meta']['note'] = (
    '7/13赎回东方AI 368.6份@4.0357(净收回1472.68→现金)，'
    '剩余368.6份(约3成仓位)。东方阿尔法6.03%持有盈亏率。'
    '永赢半导亏损-4.74%。港药回升至-0.88%。'
    '财通双基转亏。纳指+0.12%。煤炭/消费已清仓。'
)

# === Step 5: Update position_summary ===
wl['position_summary']['active_funds_count'] = len(active_funds)
wl['position_summary']['cleared_funds_count'] = len([h for h in wl['holdings'] if h['status'] == 'cleared'])
wl['position_summary']['total_active_market_value'] = round(total_holdings_mv, 2)
wl['position_summary']['cash_pct'] = wl['meta']['cash_pct']
wl['position_summary']['total_asset'] = new_total
wl['position_summary']['total_hold_return'] = round(total_hold_pnl, 2)
wl['position_summary']['total_cash'] = new_cash

sorted_active = sorted(active_funds, key=lambda h: h['market_value'], reverse=True)
top3_mv = sum(h['market_value'] for h in sorted_active[:3])
wl['position_summary']['top3_concentration_pct'] = round(top3_mv / new_total * 100, 2)

semi_codes = ['024424', '017811', '025209', '006503', '020900']
semi_mv = sum(h['market_value'] for h in active_funds if h['code'] in semi_codes)
wl['position_summary']['semiconductor_exposure_pct'] = round(semi_mv / new_total * 100, 2)
wl['position_summary']['semiconductor_exposure_pct_of_invested'] = round(semi_mv / total_holdings_mv * 100, 2)

wl['position_summary']['hk_innovation_drug_pct'] = round(
    sum(h['market_value'] for h in active_funds if h['code'] == '019671') / new_total * 100, 2)
wl['position_summary']['bond_safety_pct'] = round(
    sum(h['market_value'] for h in active_funds if h['code'] == '009022') / new_total * 100, 2)
wl['position_summary']['us_nasdaq_pct'] = round(
    sum(h['market_value'] for h in active_funds if h['code'] == '270042') / new_total * 100, 2)
wl['position_summary']['coal_defense_pct'] = 0
wl['position_summary']['consumption_pct'] = 0

wl['position_summary']['key_moves_20260713'] = (
    '1)赎回东方AI 368.6份@4.0357(净收回1472.68→现金) '
    '2)东方阿尔法MV=3180(6.03%) '
    '3)永赢半导亏损-4.74% '
    '4)港药回升至1432(-0.88%) '
    '5)财通双基转亏 '
    '6)纳指回升至110(+0.12%) '
    '7)煤炭/消费已清仓'
)

wl['position_summary']['risk_alerts'] = [
    f"半导体集中度占总资产{round(semi_mv/new_total*100,1)}% 仍偏高",
    f"东方阿尔法{sorted_active[0]['weight_pct']}%为第一大仓",
    "永赢半导亏损-4.74%(绝对敞口1785)",
    "港药回升至-0.88%但仍亏",
    f"现金{wl['meta']['cash_pct']}%大幅增加(赎回回款)，需规划再配置"
]

# Save
with open(WL_PATH, 'w', encoding='utf-8') as f:
    json.dump(wl, f, ensure_ascii=False, indent=2)

print(f"=== DONE ===")
print(f"Total: {new_total}, Cash: {new_cash} ({wl['meta']['cash_pct']}%)")
print(f"Holdings MV: {total_holdings_mv}")
print(f"Active: {len(active_funds)}, Semi: {round(semi_mv/new_total*100,1)}%")
print(f"东方AI: shares={ai['shares']}, MV={ai['market_value']}, ret={ai['hold_return_pct']}")
for h in sorted_active:
    print(f"  {h['code']} {h['name']}: MV={h['market_value']}, W={h['weight_pct']}%, ret={h['hold_return_pct']}")
